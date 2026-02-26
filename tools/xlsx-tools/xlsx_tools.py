import sys
import argparse
import json
try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

def extract_data(file_path, sheet_name=None):
    """Extracts rows from an XLSX file as JSON."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error opening XLSX: {e}")
        sys.exit(1)
        
    if sheet_name and sheet_name in wb.sheetnames:
        sheets = [wb[sheet_name]]
    else:
        sheets = wb.sheetnames
        
    output = {}
    for sheet in sheets:
        ws = wb[sheet] if isinstance(sheet, str) else sheet
        sheet_data = []
        for row in ws.iter_rows(values_only=True):
            # Filter out completely empty rows
            if any(cell is not None for cell in row):
                sheet_data.append([str(c) if c is not None else "" for c in row])
        output[ws.title] = sheet_data
        
    print(json.dumps(output, indent=2))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="XLSX Extraction Tool")
    parser.add_argument("file", help="Path to XLSX file")
    parser.add_argument("--sheet", help="Specific sheet name to extract (defaults to all)")
    
    args = parser.parse_args()
    extract_data(args.file, args.sheet)
